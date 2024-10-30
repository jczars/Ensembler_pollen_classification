#!/usr/bin/env python
# coding: utf-8

# # import

# In[1]:


import tensorflow as tf
import keras

from keras.models import Model
from keras import layers
from keras.applications.densenet import DenseNet201
import pandas as pd
import openpyxl
from tensorflow.python.keras.backend import set_session
from tensorflow.python.keras.backend import clear_session
from tensorflow.python.keras.backend import get_session
import nvidia_smi
import sys,os, gc

os.environ["tf_gpu_allocator"]="cuda_malloc_async"
sys.path.append('/media/jczars/4C22F02A22F01B22/Ensembler_pollen_classification')
print(sys.path)

from Library import utils_lib, reports_lib, models_train, sound_test_finalizado


# In[2]:
def reset_keras():
    classifier=0
    sess = get_session()
    clear_session()
    sess.close()
    sess = get_session()

    print("#"*30)

    #physical_devices = tf.config.experimental.list_physical_devices('GPU')
    #print(tf.config.experimental.set_memory_growth('/device:GPU:0', True))
    print(tf.config.experimental.get_memory_info('GPU:0'))
    print(tf.config.experimental.reset_memory_stats('GPU:0'))

    if tf.test.gpu_device_name():
        print("Default GPU Device: {}".format(tf.test.gpu_device_name()))
    else:
        print("GPU not install")

    try:
        del classifier # this is from global space - change this as you need
    except:
        pass

    print(gc.collect()) # if it's done something you should see a number being outputted
    #print(torch.cuda.empty_cache())

    # use the same config as you used to create the session
    config = tf.compat.v1.ConfigProto()
    config.gpu_options.per_process_gpu_memory_fraction = 1
    config.gpu_options.visible_device_list = "0"
    print(set_session(tf.compat.v1.Session(config=config)))
    #set_session(tf.Session(config=config))
    print("#"*30)
    print('\n')

def use_memo():
    nvidia_smi.nvmlInit()
    handle = nvidia_smi.nvmlDeviceGetHandleByIndex(0)
    info = nvidia_smi.nvmlDeviceGetMemoryInfo(handle)
    use_memo=100*info.used/info.total
    if use_memo>85:
        print('para process')
        reset_keras()
        #sys.exit()
    else:
        print('continue')

# In[3]:
    
def reports_build(test_data_generator, model, CATEGORIES, reports):
    save_dir=reports['save_dir']
    fold_var=reports['fold_var']
    batch_size=reports['batch_size']
    nm_model=reports['model']
    id_test=reports['id_test']

    (test_loss, test_accuracy) = model.evaluate(test_data_generator,
                                                batch_size=batch_size, verbose=1)
    
    
    y_true, y_pred, df_cor=reports_lib.predict_data_generator(test_data_generator,
                                                              batch_size, 
                                                              model, 
                                                              fold_var,                                                
                                                              fold_var, 
                                                              nm_model, 
                                                              CATEGORIES,
                                                              id_test, 
                                                              save_dir, 
                                                              verbose=2)

    reports_lib.plot_confusion_matrix(y_true, y_pred, CATEGORIES, nm_model, id_test,
                              save_dir, fold_var)

    #-------Rel_class
    reports_lib.class_reports(y_true, y_pred, CATEGORIES, id_test, nm_model, save_dir,
                      fold_var)

    me=reports_lib.metricas(y_true, y_pred)

    me={'test_loss':test_loss,
         'test_accuracy':test_accuracy,
         'precision':me['precision'],
         'recall':me['recall'],
         'fscore':me['fscore'],
         'kappa':me['kappa'],
         }
    reports_lib.boxplot(id_test, nm_model, df_cor, save_dir, fold_var)
    return me

# In[4]:

#alterado em 290524 às 16:13
def print_layer(conv_model, layers_params):
    save_dir=layers_params['save_dir']
    nm_model=layers_params['model']
    if not(save_dir==''):
        _csv_layers=save_dir+nm_model+'_layers.csv'
        utils_lib.add_row_csv(_csv_layers, [['id_test',layers_params['id_test']]])
        utils_lib.add_row_csv(_csv_layers, [['model',layers_params['model']]])

        utils_lib.add_row_csv(_csv_layers, [['trainable','name']])
    layers_arr=[]
    for i, layer in enumerate(conv_model.layers):
        print("{0} {1}:\t{2}".format(i, layer.trainable, layer.name))
        layers_arr.append([layer.trainable, layer.name])
    if not(save_dir==''):
        utils_lib.add_row_csv(_csv_layers, layers_arr)


# ## hyper_model

# In[9]:
    
#alterado em 290524 às 16:13
def hyper_model(rows, input_shape, num_classes, learning_rate, save_dir):
    model=eval(rows['model'])
    id_test=rows['id_test']
    # pretreined
    #image_input = Input(shape=input_shape)
    base_model = model(include_top=True,
                      weights='imagenet',
                      input_tensor=None,
                      input_shape=None,
                      pooling=None)
    #Congelar camadas convolucionais
    for layer in base_model.layers:
        layer.trainable = False
    base_model.summary()
    
    # Exclude the last 9 layers of the model.
    conv = base_model.layers[-2].output
        
    output=layers.Dense(num_classes, name='predictions', activation=rows['last_activation'])(conv)
    fine_model = Model(inputs=base_model.input, outputs=output)

    freeze = rows['freeze']
    for layer in base_model.layers[freeze:]:
        layer.trainable = True
        
    layers_params={'id_test':id_test,'save_dir':save_dir, 'model':rows['model']}
    print_layer(fine_model, layers_params)

    # otimizador
    opt=rows['optimizer']
    if opt == 'Adam':
      opt = keras.optimizers.Adam(learning_rate=learning_rate)
    if opt == 'RMSprop':
      opt=keras.optimizers.RMSprop(learning_rate=learning_rate)
    if opt == 'Adagrad':
      opt = keras.optimizers.Adagrad(learning_rate=learning_rate)
    if opt == 'SGD':
      opt = keras.optimizers.SGD(learning_rate=learning_rate)
    print(opt)

    fine_model.compile(loss="categorical_crossentropy",
                       optimizer=opt,
                       metrics=["accuracy"])
    
    fine_model.summary()
    return fine_model


# # Main


# In[11]:

def run(config,ls,_k,_p,_t):
    reset_keras()
    # ler os hyper
    nm_csv=config['path_test']+config['test_sim']
    
    print("\n[INFO] nome da planilha ",nm_csv)
    rel_book = openpyxl.load_workbook(nm_csv)

    Met_page=rel_book['Sheet']
    """
    for row in Met_page.rows:
        for cell in row:
            print(cell.value)
    """
    print(rel_book.sheetnames) # visualiza as abas existentes
    rel_book.save(nm_csv) # salva a planilha

    rows = pd.read_excel(nm_csv, sheet_name="Sheet")

    print('size _t ', _t)
    print(rows.loc[0])
    print('id_test primeira linha',rows['id_test'].loc[0])

    #constantes
    img_size = config['img_size']
    input_size=(img_size,img_size)
    input_shape = (img_size, img_size, 3)
    lr=config['learning_rate']


    for idx in range(_t):
        #alterado 300524
        reset_keras()
        use_memo()

        id=ls+idx
        row=rows.loc[id]
        print('\n---------------Teste exe')
        print(row)
        print('id_test ',row['id_test'])


        save_dir=config['path_test']
        print('save_dir ', save_dir)
        utils_lib.create_folders(save_dir, flag=0)


        #_csv_rel=save_dir+row['test_exe']
        cols_exe=['id_test', 'nm_model','k', 'val_loss','val_acc', 'test_loss', 'test_accuracy',
                   'precisio', 'recall', 'fscore', 'kappa',
                   'str_time', 'end_time', 'delay', 'epoch_finish', 'id_test', 'last_activation',
                   'activation', 'batch_size','optimizer', 'path_data', 'algoritmo']
        #utils_lib.add_row_csv(_csv_rel, cols_exe)

        _teste='Table_Exp'
        print('\n _teste ', _teste)
        if _teste in rel_book.sheetnames:
            print('exists')
        else:
            rel_book.create_sheet(_teste) # cria uma aba
        Met_page=rel_book[_teste] # inserir dados
        Met_page.append(cols_exe) # primeira linha

        for i in range(_p):
            k=_k+i
            print("\n[INFO] kfolders nº ", k)
            _path_data=row['path_data']
            #_path_data=_path_data
            print('path_data ', _path_data)
            _path_cat = _path_data+"Train/k"+str(k)
            print(_path_cat)
            CATEGORIES = sorted(os.listdir(_path_cat))
            print('classes ', CATEGORIES)
            num_classes=(len(CATEGORIES))
            print('num_classes', num_classes)

            print("\n[INFO] hyper_model ")
            model_tl=hyper_model(row, input_shape, num_classes, lr, save_dir)

            #ALTERADO COLOCAR FOR PARA DATASET

            print("\n[INFO] train, val")
            #Alterado a função para aumentar dados em tempo de execução 27/10/24
            train, val = utils_lib.load_data_train_aug(_path_data,
                              K=k,
                              BATCH=row['batch_size'],
                              INPUT_SIZE=input_size,
                              SPLIT_VALID=0.2)
            #instanciar os modelos refinados
            print("\n[INFO] start train ")
            nm_model=str(id)+'_'+row['model']
            print('id_test ',_teste, ' nm_model ', nm_model, 'k', k)

            hist, str_time, end_time, delay=models_train.run_trainV1(train,
                                                      val,
                                                      input_size,
                                                      row,
                                                      nm_model,
                                                      model_tl,
                                                      k,
                                                      config)
            print("\n[INFO] finish train ------------", delay)
            num_eapoch=len(hist.history['loss'])
            batch_size=row['batch_size']

            print("\n[INFO] memory clean ------------")
            del hist, train
            #clear_gpu()
            reset_keras()

            print("\n[INFO] evaluete ------------")
            (val_loss, val_accuracy) = model_tl.evaluate(val,
                                                           batch_size=batch_size,
                                                           verbose=1)

            test=utils_lib.load_data_test(_path_data, k, batch_size, input_size)

            #build 310524
            reports={'model':row['model'],'batch_size':row['batch_size'],
                     'save_dir':save_dir, 'fold_var':k, 'id_test':row['id_test']}
            me=reports_build(test, model_tl, CATEGORIES, reports)


            print("\n[INFO] save evaluete ------------")
            #buil 290524 às 18:35
            prec=3
            va=round(val_accuracy, prec)
            ta=round(me['test_accuracy'], prec)


            data=[str(id), row['model'], k, val_loss, va,
                   me['test_loss'], ta,
                   me['precision'], me['recall'], me['fscore'], me['kappa'],
                   str_time, end_time, delay, num_eapoch, row['id_test'],
                   row['last_activation'], row['activation'],row['batch_size'], 
                   row['optimizer'], row['path_data'], config['algoritmo']]
            #utils_lib.add_row_csv(_csv_rel, data)

            Met_page.append(data)
            rel_book.save(nm_csv)

            #limpar memoria
            print("\n[INFO] memory clean ------------")
            del _path_data, _path_cat, CATEGORIES, num_classes
            del data, test, me
            del model_tl, str_time, end_time, delay, val
            
            del reports
            
            del val_loss, val_accuracy
            del batch_size, num_eapoch

            reset_keras()

if __name__=="__main__":
    # Sets the working directory
    os.chdir('/media/jczars/4C22F02A22F01B22/Ensembler_pollen_classification/')
    # Obtém o nome do script em execução
    script_name = os.path.basename(__file__)
    print(f" script_name {script_name} ")
    
    
    config={'epochs':500,
            'alpha': 1e-5,
             'learning_rate':1e-4,
             'img_size': 224,
             'path_test': '1_Fine_tuned/Reports_DFT_Dn_281024/',
             'test_sim': '0_CLASS_Dn201_DFT_VISTAS_lr.xlsx',
             'algoritmo':script_name
             }
    ls=0
    _k=1
    _p=1
    _t=1
    
    run(config,ls,_k,_p,_t)
    sound_test_finalizado.beep(2)