#!/usr/bin/env python
# coding: utf-8

# # import

# In[1]:


import tensorflow as tf
import keras
import os
os.environ["tf_gpu_allocator"]="cuda_malloc_async"


# In[2]:

from keras.models import Model
from keras import layers
from keras.applications.densenet import DenseNet201
#from keras.applications import VGG16
#from keras.layers import Dense, Flatten, Dropout, BatchNormalization, Dropout
#from keras.layers import Conv2D, GlobalAveragePooling2D, Reshape
from keras.callbacks import EarlyStopping, ModelCheckpoint, ReduceLROnPlateau
import datetime

import numpy as np
import pandas as pd
#from tqdm import tqdm
#import glob
from sklearn.metrics import confusion_matrix, classification_report
import openpyxl
#from openpyxl.utils.dataframe import dataframe_to_rows
import matplotlib.pyplot as plt
import seaborn as sns
from sklearn import metrics
from tensorflow.keras.preprocessing.image import ImageDataGenerator
from tensorflow.python.keras.backend import set_session
from tensorflow.python.keras.backend import clear_session
from tensorflow.python.keras.backend import get_session
import nvidia_smi

#090624
#import openpyxl
#from openpyxl.utils.dataframe import dataframe_to_rows

# In[3]:


import sys,os, gc
sys.path.append('/media/jczars/4C22F02A22F01B22/Ensembler_pollen_classification')
print(sys.path)
#from Library import reports_lib_xlx as reports_lib
from Library import utils_lib,  sound_test_finalizado
#from Library import send_whatsApp_msn

def use_memo():
    nvidia_smi.nvmlInit()
    handle = nvidia_smi.nvmlDeviceGetHandleByIndex(0)
    info = nvidia_smi.nvmlDeviceGetMemoryInfo(handle)
    use_memo=100*info.used/info.total
    if use_memo>85:
        print('para process')
        reset_keras()
        #sys.exit()
    else:
        print('continue')


# # Functions
def predict_data_generator(test_data_generator, batch_size, model, k, fold_var,
                             nm_model, CATEGORIES, id_test, save_dir='', verbose=2):

    """
    --> evaluation metrics (accuracy_score, precision, recall, fscore, kappa)
    --Version V1 17/04/24 09:59
    :param: test_data_generator: images from Image Date Generator
    :param: batch_size: labels predict
    :param: model: model load
    :param: save_dir: model load
    :param: k: number k in kfolders
    :param: CATEGORIES: list classes the of images
    :param: verbose: enable printing
    :return: print classification reports
    """
    filenames = test_data_generator.filenames
    y_true = test_data_generator.classes
    df = pd.DataFrame(filenames, columns=['filenames'])
    confianca = []
    nb_samples = len(filenames)
    y_preds = model.predict(test_data_generator,
                            steps=nb_samples // batch_size + 1)

    for i in range(len(y_preds)):
        confi = np.max(y_preds[i])
        confianca.append(confi)
        if verbose == 1:
            print('i ', i, ' ', y_preds[i])
            print('confi ', confi)

    y_pred = np.argmax(y_preds, axis=1)
    if verbose == 2:
        print('Size y_true', len(y_true))
        print('Size y_pred', len(y_pred))

    df['y_true'] = y_true
    df['y_pred'] = y_pred
    df['confianca'] = confianca
    df.insert(loc=2, column='labels', value='')
    df.insert(loc=4, column='predict', value='')
    df.insert(loc=6, column='sit', value='')

    # veficar acertos
    for idx, row in df.iterrows():
        cat_true = CATEGORIES[row['y_true']]
        cat_pred = CATEGORIES[row['y_pred']]
        if verbose == 1:
            print('cat_true ', cat_true, 'cat_pred ', cat_pred)
        df.at[idx, 'labels'] = cat_true
        df.at[idx, 'predict'] = cat_pred
        if row['y_true'] == row['y_pred']:
            df.at[idx, 'sit'] = 'C'
        else:
            df.at[idx, 'sit'] = 'E'

    df = df.sort_values(by='labels')
    df_Err = df[df['sit'] == 'E']
    df_cor = df[df['sit'] == 'C']


    if not (save_dir == ''):
        df_Err.to_csv(save_dir+'filterWrong_'+str(id_test)+'_'+nm_model+'_k'+
                    str(fold_var) +'.csv', index=True)

        df_cor.to_csv(save_dir+'filterCorrect_'+str(id_test)+'_'+nm_model+'_k'+
                    str(fold_var) +'.csv', index=True)



    return y_true, y_pred, df_cor

def plot_confusion_matrix(y_true, y_pred, CATEGORIES, nm_model, id_test,
                          save_dir='', fold_var=0, verbose=0):

    """
  --> Confusion Matrix, print or save
  :param: test_data_generator: images from Image Date Generator
  :param: y_pred: labels predict
  :param: CATEGORIES: list classes the of images
  :param: save_dir: path to save confusion Matrix, standar ='', not save
  :param: fold_var: numers order
  :param: verbose: if == 1, print graph the confusion matrix, if == 0, not print
  :return: vector of the metrics
  """
    print('Confusion Matrix')
    mat = confusion_matrix(y_true, y_pred,
                           normalize=None)
    # Save matrix 21/01/24
    print('salvando matriz como csv em ', save_dir)
    df_mat = pd.DataFrame(mat)

    if not (save_dir == ''):
        df_mat.to_csv(save_dir +'mat_conf_test_'+str(id_test)+'_'+nm_model+'_k'+
                    str(fold_var) + '.csv')

    #plt.figure(figsize=fig_size)
    my_dpi=100
    plt.figure(figsize=(900/my_dpi, 900/my_dpi), dpi=my_dpi)
    ax = plt.subplot()
    sns.heatmap(mat, cmap="Blues", annot=True)  # annot=True to annotate cells

    ax.set_xticklabels(ax.get_xticklabels(),
                       rotation=90, horizontalalignment='right', fontsize=8)
    ax.set_yticklabels(ax.get_yticklabels(),
                       rotation=0, horizontalalignment='right', fontsize=8)

    # labels, title and ticks
    ax.set_xlabel('Rótulos previstos')
    ax.set_ylabel('Rótulos verdadeiros')
    ax.set_title('Matriz de confusão')

    if len(CATEGORIES)<21:
        ax.xaxis.set_ticklabels(CATEGORIES)
        ax.yaxis.set_ticklabels(CATEGORIES)


    if not (save_dir == ''):
        #plt.rcParams['savefig.dpi'] = 300
        #dpi=300
        plt.savefig(save_dir +'mat_conf_test_'+str(id_test)+'_'+nm_model+'_k'+
                    str(fold_var) + '.jpg')
    if verbose == 1:
        plt.show()

def metricas(y_true, y_pred):
      """
      --> evaluation metrics (accuracy_score, precision, recall, fscore, kappa)
    :param: y_true: labels real
    :param: y_pred: labels predict
    :return: vector of the metrics
    """
      #-------metrics
      print('\n3-Metrics')
      print('Acc, precision, recall, fscore, kappa')
      #me=reports_lib.metricasV1(y_true, y_pred)

      precision, recall, fscore, support = metrics.precision_recall_fscore_support(
      y_true, y_pred)
      kappa = metrics.cohen_kappa_score(y_true, y_pred)
      prec=4

      #alterado 090624
      _p=round(np.mean(precision, axis=0), prec)
      _r=round(np.mean(recall, axis=0), prec)
      _f=round(np.mean(fscore, axis=0), prec)
      _kp=round(kappa, prec)
      me ={'precision':_p,
       'recall':_r,
       'fscore':_f,
       'kappa':_kp}

      return me

def class_reports(y_true, y_pred, CATEGORIES, id_test, nm_model, save_dir='',
                  fold_var=0):
    """
  --> evaluation metrics (accuracy_score, precision, recall, fscore, kappa)
  --Version V1 17/04/24 09:59
  :param: y_true: labels real
  :param: y_pred: labels predict
  :param: CATEGORIES: list classes the of images
  :return: print classification reports
  """

    print('Classification Report')
    print(classification_report(y_true, y_pred, target_names=CATEGORIES))
    # Salva o relatório de classificação
    report = classification_report(
        y_true, y_pred, target_names=CATEGORIES, output_dict=True)
    df_report = pd.DataFrame(report).transpose()
    if not (save_dir == ''):
        df_report.to_csv(save_dir+'class_report_test_'+str(id_test)+'_'+nm_model+
                     '_k'+str(fold_var) + '.csv', index=True)

def boxplot(id_test, nm_model, df_corre, save_dir='', fold_var=0):
    my_dpi=100
    plt.figure(figsize=(900/my_dpi, 900/my_dpi), dpi=my_dpi)

    sns.set_style("whitegrid")

    # Adicionando Título ao gráfico
    sns.boxplot(y=df_corre["labels"], x=df_corre["confianca"])
    plt.title("Classes sem erros de classificação", loc="center", fontsize=18)
    plt.xlabel("acurácia")
    plt.ylabel("classes")

    if not (save_dir == ''):
        plt.savefig(save_dir+'/boxplot_correct_'+str(id_test)+'_'+nm_model+'_k'+
                    str(fold_var) + '.jpg')
    plt.show()

def reports_build(test_data_generator, model, CATEGORIES, reports):
    save_dir=reports['save_dir']
    fold_var=reports['fold_var']
    batch_size=reports['batch_size']
    nm_model=reports['model']
    id_test=reports['id_test']

    (test_loss, test_accuracy) = model.evaluate(test_data_generator,
                                                batch_size=batch_size, verbose=1)

    y_true, y_pred, df_cor=predict_data_generator(test_data_generator,
                                                  batch_size, model, fold_var,
                                                  fold_var, nm_model, CATEGORIES,
                                                  id_test, save_dir, verbose=2)

    plot_confusion_matrix(y_true, y_pred, CATEGORIES, nm_model, id_test,
                              save_dir, fold_var)

    #-------Rel_class
    class_reports(y_true, y_pred, CATEGORIES, id_test, nm_model, save_dir,
                      fold_var)

    me=metricas(y_true, y_pred)

    me={'test_loss':test_loss,
         'test_accuracy':test_accuracy,
         'precision':me['precision'],
         'recall':me['recall'],
         'fscore':me['fscore'],
         'kappa':me['kappa'],
         }
    boxplot(id_test, nm_model, df_cor, save_dir, fold_var)
    return me



# ## print_layer_trainable

# In[4]:

#alterado em 290524 às 16:13
def print_layer(conv_model, layers_params):
    save_dir=layers_params['save_dir']
    nm_model=layers_params['model']
    if not(save_dir==''):
        _csv_layers=save_dir+nm_model+'_layers.csv'
        utils_lib.add_row_csv(_csv_layers, [['id_test',layers_params['id_test']]])
        utils_lib.add_row_csv(_csv_layers, [['model',layers_params['model']]])

        utils_lib.add_row_csv(_csv_layers, [['trainable','name']])
    layers_arr=[]
    for i, layer in enumerate(conv_model.layers):
        print("{0} {1}:\t{2}".format(i, layer.trainable, layer.name))
        layers_arr.append([layer.trainable, layer.name])
    if not(save_dir==''):
        utils_lib.add_row_csv(_csv_layers, layers_arr)


# ## callbacks

# In[5]:


def get_model_name(nm_modelo, k):
    return nm_modelo+'_bestLoss_' + str(k) + '.keras'


# In[6]:


def callbacks(save_dir, nm_modelo, k, alpha):
    cp_loss = ModelCheckpoint(save_dir + get_model_name(nm_modelo, k),
                                  monitor='val_loss',
                                  verbose=1,
                                  save_best_only=True,
                                  mode='min')
    es = EarlyStopping(monitor='val_loss',
               mode='min',
               patience = 10,
               verbose=1)
    lr = ReduceLROnPlateau(monitor='val_loss',
                   factor=0.1,
                   min_delta=alpha,
                   patience=5,
                   verbose=1)
    callbacks_list = [cp_loss, es]
    return callbacks_list


# ## clear_gpu

# In[7]:


def reset_keras():
    classifier=0
    sess = get_session()
    clear_session()
    sess.close()
    sess = get_session()

    print("#"*30)

    #physical_devices = tf.config.experimental.list_physical_devices('GPU')
    #print(tf.config.experimental.set_memory_growth('/device:GPU:0', True))
    print(tf.config.experimental.get_memory_info('GPU:0'))
    print(tf.config.experimental.reset_memory_stats('GPU:0'))

    if tf.test.gpu_device_name():
        print("Default GPU Device: {}".format(tf.test.gpu_device_name()))
    else:
        print("GPU not install")

    try:
        del classifier # this is from global space - change this as you need
    except:
        pass

    print(gc.collect()) # if it's done something you should see a number being outputted
    #print(torch.cuda.empty_cache())

    # use the same config as you used to create the session
    config = tf.compat.v1.ConfigProto()
    config.gpu_options.per_process_gpu_memory_fraction = 1
    config.gpu_options.visible_device_list = "0"
    print(set_session(tf.compat.v1.Session(config=config)))
    #set_session(tf.Session(config=config))
    print("#"*30)
    print('\n')


# ## run_train

# In[8]:


def run_train(train, val, input_size, rows,nm_model,
              model_fine, k, config):

    print('batch_size ', rows['batch_size'])
    callbacks_list=callbacks(config['path_test'], nm_model, k, config['alpha'])
    str_time = datetime.datetime.now().replace(microsecond=0)

    with tf.device('/device:GPU:0'):
        print('\n',str_time)
        hist = model_fine.fit(train,
                              batch_size=rows['batch_size'],
                              epochs=config['epochs'],
                              callbacks=callbacks_list,
                              verbose=1,
                              validation_data=val)
    end_time = datetime.datetime.now().replace(microsecond=0)
    delay=end_time-str_time
    print('Training time: %s' % (delay))

    return hist, str_time, end_time, delay


# ## hyper_model

# In[9]:
    
def load_data_train(PATH_BD, K, BATCH, INPUT_SIZE, SPLIT_VALID):
    """
    -->loading train data 
    :param: PATH_BD: file name 
    :param: K: k the kfolders values
    :param: BATCH: batch size
    :param: INPUT_SIZE: input dimensions, height and width, default=(224,224)
    :param: SPLIT_VALID: portion to divide the training base into training and validation
    return: train and valid dataset
    """
    train_dir = PATH_BD + '/Train/k' + str(K)
    print('train_dir ', train_dir)
    
    #idg = ImageDataGenerator(rescale=1. / 255, validation_split=SPLIT_VALID)
    
    idg = ImageDataGenerator(width_shift_range=0.1, 
                             height_shift_range=0.1,
                             zoom_range=0.3, 
                             fill_mode='nearest', 
                             horizontal_flip=True, 
                             rescale=1./255,
                             validation_split=SPLIT_VALID) # set validation split

    train_generator = idg.flow_from_directory(
        directory=train_dir,
        target_size=INPUT_SIZE,
        color_mode="rgb",
        batch_size=BATCH,
        class_mode="categorical",
        shuffle=True,
        seed=42,
        subset='training')

    val_generator = idg.flow_from_directory(
        directory=train_dir,
        target_size=INPUT_SIZE,
        color_mode="rgb",
        batch_size=BATCH,
        class_mode="categorical",
        shuffle=True,
        seed=42,
        subset='validation')

    return train_generator, val_generator

#alterado em 290524 às 16:13
def hyper_model(rows, input_shape, num_classes, learning_rate, save_dir):
    model=eval(rows['model'])
    id_test=rows['id_test']
    # pretreined
    #image_input = Input(shape=input_shape)
    base_model = model(include_top=True,
                      weights='imagenet',
                      input_tensor=None,
                      input_shape=None,
                      pooling=None)
    #Congelar camadas convolucionais
    for layer in base_model.layers:
        layer.trainable = False
    base_model.summary()
    
    # Exclude the last 9 layers of the model.
    conv = base_model.layers[-2].output
        
    output=layers.Dense(num_classes, name='predictions', activation=rows['last_activation'])(conv)
    fine_model = Model(inputs=base_model.input, outputs=output)

    freeze = rows['freeze']
    for layer in base_model.layers[freeze:]:
        layer.trainable = True
        
    layers_params={'id_test':id_test,'save_dir':save_dir, 'model':rows['model']}
    print_layer(fine_model, layers_params)

    # otimizador
    opt=rows['optimizer']
    if opt == 'Adam':
      opt = keras.optimizers.Adam(learning_rate=learning_rate)
    if opt == 'RMSprop':
      opt=keras.optimizers.RMSprop(learning_rate=learning_rate)
    if opt == 'Adagrad':
      opt = keras.optimizers.Adagrad(learning_rate=learning_rate)
    if opt == 'SGD':
      opt = keras.optimizers.SGD(learning_rate=learning_rate)
    print(opt)

    fine_model.compile(loss="categorical_crossentropy",
                       optimizer=opt,
                       metrics=["accuracy"])
    
    fine_model.summary()
    return fine_model


# # Main


# In[11]:

def run(config,ls,_k,_p,_t):
    reset_keras()
    # ler os hyper
    nm_csv=config['path_test']+config['test_sim']
    """
    Testes = pd.read_csv(nm_csv)
    total_test=len(Testes)
    print(Testes.head())
    """

    print("\n[INFO] nome da planilha ",nm_csv)
    rel_book = openpyxl.load_workbook(nm_csv)

    Met_page=rel_book['Sheet']
    """
    for row in Met_page.rows:
        for cell in row:
            print(cell.value)
    """
    print(rel_book.sheetnames) # visualiza as abas existentes
    rel_book.save(nm_csv) # salva a planilha

    rows = pd.read_excel(nm_csv, sheet_name="Sheet")

    print('size _t ', _t)
    print(rows.loc[0])
    print('id_test primeira linha',rows['id_test'].loc[0])

    #constantes
    img_size = config['img_size']
    input_size=(img_size,img_size)
    input_shape = (img_size, img_size, 3)
    lr=config['learning_rate']


    for idx in range(_t):
        #alterado 300524
        reset_keras()
        use_memo()

        id=ls+idx
        row=rows.loc[id]
        print('\n---------------Teste exe')
        print(row)
        print('id_test ',row['id_test'])


        save_dir=config['path_test']
        print('save_dir ', save_dir)
        utils_lib.create_folders(save_dir, flag=0)


        #_csv_rel=save_dir+row['test_exe']
        cols_exe=['id_test', 'nm_model','k', 'val_loss','val_acc', 'test_loss', 'test_accuracy',
                   'precisio', 'recall', 'fscore', 'kappa',
                   'str_time', 'end_time', 'delay', 'epoch_finish', 'id_test', 'last_activation',
                   'activation', 'batch_size','optimizer', 'path_data', 'algoritmo']
        #utils_lib.add_row_csv(_csv_rel, cols_exe)

        _teste='Table_Exp'
        print('\n _teste ', _teste)
        if _teste in rel_book.sheetnames:
            print('exists')
        else:
            rel_book.create_sheet(_teste) # cria uma aba
        Met_page=rel_book[_teste] # inserir dados
        Met_page.append(cols_exe) # primeira linha

        for i in range(_p):
            k=_k+i
            print("\n[INFO] kfolders nº ", k)
            _path_data=row['path_data']
            #_path_data=_path_data
            print('path_data ', _path_data)
            _path_cat = _path_data+"Train/k"+str(k)
            print(_path_cat)
            CATEGORIES = sorted(os.listdir(_path_cat))
            print('classes ', CATEGORIES)
            num_classes=(len(CATEGORIES))
            print('num_classes', num_classes)

            print("\n[INFO] hyper_model ")
            model_tl=hyper_model(row, input_shape, num_classes, lr, save_dir)

            #ALTERADO COLOCAR FOR PARA DATASET

            print("\n[INFO] train, val")
            #Alterado a função para aumentar dados em tempo de execução 27/10/24
            train, val = load_data_train(_path_data,
                              K=k,
                              BATCH=row['batch_size'],
                              INPUT_SIZE=input_size,
                              SPLIT_VALID=0.2)
            #instanciar os modelos refinados
            print("\n[INFO] start train ")
            nm_model=str(id)+'_'+row['model']
            print('id_test ',_teste, ' nm_model ', nm_model, 'k', k)

            hist, str_time, end_time, delay=run_train(train,
                                                      val,
                                                      input_size,
                                                      row,
                                                      nm_model,
                                                      model_tl,
                                                      k,
                                                      config)
            print("\n[INFO] finish train ------------", delay)
            num_eapoch=len(hist.history['loss'])
            batch_size=row['batch_size']

            # 4-plot Accuracy
            pd_history = pd.DataFrame(hist.history)
            pd_history.plot()
            plt.grid(True)
            plt.savefig(save_dir + 'acc_loss_' +str(row['id_test'])+'_'+nm_model+'_k'+ str(k) + '.jpg')
            plt.show()

            print("\n[INFO] memory clean ------------")
            del pd_history
            del hist, train
            #clear_gpu()
            reset_keras()

            print("\n[INFO] evaluete ------------")
            (val_loss, val_accuracy) = model_tl.evaluate(val,
                                                           batch_size=batch_size,
                                                           verbose=1)

            test=utils_lib.load_data_test(_path_data, k, batch_size, input_size)

            #build 310524
            reports={'model':row['model'],'batch_size':row['batch_size'],
                     'save_dir':save_dir, 'fold_var':k, 'id_test':row['id_test']}
            me=reports_build(test, model_tl, CATEGORIES, reports)


            print("\n[INFO] save evaluete ------------")
            #buil 290524 às 18:35
            prec=3
            va=round(val_accuracy, prec)
            ta=round(me['test_accuracy'], prec)


            data=[str(id), row['model'], k, val_loss, va,
                   me['test_loss'], ta,
                   me['precision'], me['recall'], me['fscore'], me['kappa'],
                   str_time, end_time, delay, num_eapoch, row['id_test'],
                   row['last_activation'], row['activation'],row['batch_size'], 
                   row['optimizer'], row['path_data'], config['algoritmo']]
            #utils_lib.add_row_csv(_csv_rel, data)

            Met_page.append(data)
            rel_book.save(nm_csv)

            #limpar memoria
            print("\n[INFO] memory clean ------------")
            del _path_data, _path_cat, CATEGORIES, num_classes
            del data, test, me
            del model_tl, str_time, end_time, delay, val
            
            del reports
            
            del val_loss, val_accuracy
            del batch_size, num_eapoch

            reset_keras()

if __name__=="__main__":
    # Sets the working directory
    os.chdir('/media/jczars/4C22F02A22F01B22/Ensembler_pollen_classification/')
    
    config={'epochs':500,
            'alpha': 1e-5,
             'learning_rate':1e-4,
             'img_size': 224,
             'path_test': '1_Fine_tuned/Reports_DFT_Dn_lr_281024/',
             'test_sim': '0_CLASS_Dn201_DFT_VISTAS_lr.xlsx',
             'algoritmo':'1_Fine_tuned/OrigModel_DFT_K10_RC_xlsx_271024.py'
             }
    ls=1
    _k=8
    _p=3
    _t=1
    
    run(config,ls,_k,_p,_t)
    sound_test_finalizado.beep(5)