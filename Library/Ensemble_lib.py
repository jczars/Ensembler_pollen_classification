#!/usr/bin/env python
# coding: utf-8

"""
Ensemble Prediction Script
==========================
This script is designed for ensemble predictions in a pollen classification context using TensorFlow.
It includes functions for generating predictions, saving results to Excel, and performing model evaluations.

Requirements:
-------------
- TensorFlow
- Pandas
- Openpyxl
- Matplotlib

Usage:
------
Set the `test_data_directory`, `model_directory`, `output_directory`, and specify parameters as needed.
"""



import numpy as np
import pandas as pd
#from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Utility Imports
# Ensure the following path includes the required libraries: utils_lib, reports_lib, models_train, cam
# sys.path.append('/path/to/your/custom/libraries')
#from Library import utils_lib, reports_lib, models_train, grad_cam_lib as cam


def predict_data_generator(test_data_generator, model, categories, batch_size=32, k_folds=None, verbose=1):
    """
    Generate predictions and evaluate model performance on a test data generator.

    Parameters:
    -----------
    test_data_generator : tensorflow.keras.preprocessing.image.ImageDataGenerator
        Generator yielding test images for predictions.
    model : tf.keras.Model
        Trained model used to generate predictions.
    categories : list of str
        List of class names corresponding to model output.
    batch_size : int, optional
        Number of samples per gradient update (default is 32).
    k_folds : int, optional
        Number of folds for k-fold validation if applicable.
    verbose : int, optional
        Verbosity level. 0 = silent, 1 = summary info, 2 = detailed info.

    Returns:
    --------
    dict
        Dictionary containing true labels, predicted labels, and confidence scores.
    """
    filenames = test_data_generator.filenames
    y_true = test_data_generator.classes
    nb_samples = len(filenames)

    # DataFrame to track filenames and predictions
    df_results = pd.DataFrame(filenames, columns=['filenames'])
    confidence_scores = []

    # Generate predictions
    y_preds = model.predict(test_data_generator, steps=np.ceil(nb_samples / batch_size), verbose=verbose)
    
    # Calculate confidence scores and predicted classes
    for pred in y_preds:
        confidence_scores.append(np.max(pred))  # Max confidence score per prediction
    
    y_pred = np.argmax(y_preds, axis=1)

    # Verbose output
    if verbose >= 1:
        print(f"Number of Samples: {nb_samples}")
        print(f"True Labels Size: {len(y_true)}, Predicted Labels Size: {len(y_pred)}")
    if verbose >= 2:
        for idx, (pred, conf) in enumerate(zip(y_preds, confidence_scores)):
            print(f"Sample {idx} Prediction: {pred} - Confidence: {conf}")

    # Append results to DataFrame
    df_results['y_true'] = y_true
    df_results['y_pred'] = y_pred
    df_results['confidence'] = confidence_scores

    # Return organized results in dictionary format
    return {
        "y_true": y_true,
        "y_pred": y_pred,
        "confidence_scores": confidence_scores,
        "results_df": df_results
    }

def ensemble_predict(
    test, batch_size, en_models, save_dir, k, categories, rel_book, nm_rel
):
    """
    Realiza previsões de classificação em conjunto usando múltiplos modelos e armazena os resultados em um DataFrame,
    além de salvar os resultados em uma planilha Excel.

    Parâmetros:
    -----------
    test : keras.preprocessing.image.DirectoryIterator
        O conjunto de dados de teste, contendo imagens e rótulos.
    batch_size : int
        O tamanho do lote para a predição.
    en_models : list
        Lista de modelos de machine learning para usar na predição em conjunto.
    save_dir : str
        Caminho do diretório onde resultados intermediários podem ser salvos.
    k : int
        Identificador para diferenciar execuções, usado no nome da aba no Excel.
    categories : list
        Lista de categorias de classificação.
    rel_book : openpyxl.Workbook
        Workbook do Excel onde os resultados serão salvos.
    nm_rel : str
        Nome do arquivo onde a planilha será salva.

    Retorna:
    --------
    pd.DataFrame
        Um DataFrame contendo as previsões e as confianças dos modelos para cada imagem.
    """
    filenames = test.filenames
    results_df = pd.DataFrame(filenames, columns=["filenames"])

    # Realiza predições com cada modelo do conjunto
    for i, model in enumerate(en_models):
        y_true, y_pred, confidence = predict_data_generator(
            test, batch_size, model, save_dir, k, categories, verbose=1
        )

        # Armazena os resultados no DataFrame com colunas únicas para cada modelo
        results_df[f"y_true_{i}"] = y_true
        results_df[f"y_pred_{i}"] = y_pred
        results_df[f"confidence_{i}"] = confidence

    # Nome da aba no Excel
    sheet_name = f"TableClassification_{k}"
    save_results_to_excel(results_df, rel_book, sheet_name, nm_rel)

    return results_df


def save_results_to_excel(df, workbook, sheet_name, file_name):
    """
    Função auxiliar para salvar um DataFrame em uma nova aba de um Workbook do Excel.

    Parâmetros:
    -----------
    df : pd.DataFrame
        DataFrame contendo os dados a serem salvos.
    workbook : openpyxl.Workbook
        Workbook do Excel onde os dados serão adicionados.
    sheet_name : str
        Nome da aba a ser criada no Workbook.
    file_name : str
        Nome do arquivo onde a planilha será salva.
    """
    # Cria uma nova aba no Workbook
    workbook.create_sheet(sheet_name)
    page = workbook[sheet_name]

    # Adiciona os dados do DataFrame na aba
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            page.cell(row=r_idx, column=c_idx, value=value)

    # Salva o Workbook no arquivo especificado
    workbook.save(file_name)
    
def verifica_votos(df, rel_book, nm_rel, verbose=0):
    """
    Verifica as previsões dos modelos em um DataFrame de votação, onde cada linha representa uma predição e seu respectivo rótulo.
    Atualiza um DataFrame de votos com a predição final e escreve os resultados em uma planilha Excel.

    Parâmetros:
    -----------
    df : pd.DataFrame
        DataFrame contendo os dados de previsões dos modelos (colunas: 'filenames', 'y_true0', 'y_pred0', 'y_pred1', 'conf0', 'conf1').
    rel_book : openpyxl.Workbook
        Workbook onde os resultados serão salvos.
    nm_rel : str
        Nome do arquivo Excel onde o workbook será salvo.
    verbose : int, opcional
        Nível de detalhamento das mensagens de debug. 0 para desativado, 2 para detalhamento completo.

    Retorna:
    --------
    pd.DataFrame
        DataFrame contendo as previsões finais, confiança e situação ('C' para correto e 'E' para erro).
    """
    # Inicializa o DataFrame de votos com colunas específicas
    votos = pd.DataFrame({
        'filenames': df['filenames'],
        'y_true': df['y_true0'],
        'y_pred': '',
        'conf': '',
        'sit': ''
    })

    # Processa cada linha do DataFrame para verificar a votação entre os modelos
    for i in range(len(df)):
        y_pred0, y_pred1 = df.loc[i, ['y_pred0', 'y_pred1']]
        y_true = df.loc[i, 'y_true0']
        conf = [df.loc[i, 'conf0'], df.loc[i, 'conf1']]

        if y_pred0 == y_pred1:
            pred = y_pred0
            max_conf = conf[0] if conf[0] >= conf[1] else conf[1]
        else:
            maxIdx = np.argmax(conf)
            pred = df.loc[i, f'y_pred{maxIdx}']
            max_conf = conf[maxIdx]

        # Atualiza o DataFrame de votos
        votos.at[i, 'y_pred'] = pred
        votos.at[i, 'conf'] = max_conf
        votos.at[i, 'sit'] = 'C' if y_true == pred else 'E'

        # Exibe detalhes se verbose estiver configurado
        if verbose == 2:
            print(
                f"\n{'verdade' if y_pred0 == y_pred1 else 'falso'}\n"
                f"ord {i}, y_pred0 {y_pred0}, y_pred1 {y_pred1}, conf {conf}\n"
                f"maxIdx {np.argmax(conf)}, conf max {max_conf}, pred {pred}"
            )

    # Nome da aba no Excel e chamada para salvar o DataFrame de votos
    sheet_name = "EnsembleVote"
    save_results_to_excel(votos, rel_book, sheet_name, nm_rel, ['filenames', 'y_true', 'y_pred', 'conf', 'sit'])

    # Calcula e exibe total de erros
    total_erros = votos['sit'].value_counts().get('E', 0)
    total = len(votos)
    print(f'Total de erros: {total_erros}')
    print(f'Precisão: {1 - total_erros / total:.2%}')

    return votos


def save_results_to_excel(df, workbook, sheet_name, file_name, columns):
    """
    Função auxiliar para salvar um DataFrame em uma nova aba de um Workbook do Excel.

    Parâmetros:
    -----------
    df : pd.DataFrame
        DataFrame contendo os dados a serem salvos.
    workbook : openpyxl.Workbook
        Workbook do Excel onde os dados serão adicionados.
    sheet_name : str
        Nome da aba a ser criada no Workbook.
    file_name : str
        Nome do arquivo onde a planilha será salva.
    columns : list
        Lista de colunas para incluir no cabeçalho da aba.
    """
    workbook.create_sheet(sheet_name)
    page = workbook[sheet_name]

    # Adiciona o cabeçalho
    page.append(columns)

    # Adiciona os dados do DataFrame na aba
    for r_idx, row in enumerate(dataframe_to_rows(df[columns], index=False, header=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            page.cell(row=r_idx, column=c_idx, value=value)

    # Salva o Workbook no arquivo especificado
    workbook.save(file_name)
    
